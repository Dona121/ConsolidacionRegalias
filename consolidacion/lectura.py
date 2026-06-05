"""
Funciones de lectura y normalización de los archivos cargados.

- ruta_reciente        : escoge el archivo más nuevo dentro de un prefijo.
- leer_excel_regalias  : lee un reporte Gesproy respetando su metadata.
- leer_tabla_excel     : lee una tabla con nombre. Usa polars si soporta
                         table_name; si no, hace fallback a openpyxl.
- normalizar_fecha     : convierte columnas a pl.Date independientemente
                         del dtype con que Polars las haya leído.
"""

from __future__ import annotations

import io
import datetime as _dt

import polars as pl

from .validacion import dtype_es_fecha


_TIPOS_NUMERICOS = {
    pl.Int8, pl.Int16, pl.Int32, pl.Int64,
    pl.UInt8, pl.UInt16, pl.UInt32, pl.UInt64,
    pl.Float32, pl.Float64,
}


def ruta_reciente(archivos: list[str], prefijo: str) -> str | None:
    """
    Devuelve el nombre del archivo cuyo nombre comienza con `prefijo` y
    tiene la fecha (YYYYMMDD) más reciente. Si ningún archivo tiene fecha en
    el nombre, devuelve el primer match.
    """
    filtrados = [f for f in archivos if f.startswith(prefijo)]
    if not filtrados:
        return None
    filas = []
    for nombre in filtrados:
        fecha = pl.Series([nombre]).str.extract(r"(\d{8})").item()
        if fecha is not None:
            filas.append({"Nombre archivo": nombre, "Fecha": fecha})
    if not filas:
        return filtrados[0]
    return (
        pl.DataFrame(filas)
        .sort("Fecha", descending=True)
        .row(0, named=True)["Nombre archivo"]
    )


def leer_excel_regalias(contenido_bytes: bytes) -> pl.DataFrame:
    """
    Lee un reporte Gesproy donde la fila 0 es metadato, la fila 1 son los
    encabezados y los datos comienzan en la fila 2. Todas las columnas se
    leen como texto (infer_schema_length=0).
    """
    df = pl.read_excel(
        io.BytesIO(contenido_bytes),
        has_header=False,
        infer_schema_length=0,
    )
    if df.height < 2:
        raise ValueError(
            "El archivo no tiene suficientes filas. "
            "Se esperan al menos 2 (fila de metadato + fila de encabezados)."
        )
    encabezados = dict(zip(df.columns, df.row(1)))
    return (
        df
        .rename(encabezados)
        .slice(2)
        .select(pl.all().name.map(lambda x: x.strip()))
    )


def leer_tabla_excel(contenido_bytes: bytes, nombre_tabla: str) -> pl.DataFrame:
    """
    Lee una tabla Excel con nombre (objeto Tabla, no solo una hoja).

    Estrategia en dos pasos:

      1) Intenta pl.read_excel(..., table_name=...). Requiere polars 1.0+
         con fastexcel instalado.
      2) Si esa firma no existe (polars antiguo), hace fallback a openpyxl:
         busca la tabla por nombre en todas las hojas, calcula su rango y
         devuelve un DataFrame con los valores.

    Lanza ValueError si la tabla no aparece en ninguna hoja.
    """
    df = None
    try:
        df = pl.read_excel(
            io.BytesIO(contenido_bytes),
            table_name=nombre_tabla,
        )
    except TypeError as e:
        if "table_name" not in str(e):
            raise
    except Exception:
        pass

    if df is None:
        df = _leer_tabla_openpyxl(contenido_bytes, nombre_tabla)

    return _normalizar_nombres_columnas(df)


def _normalizar_nombres_columnas(df: pl.DataFrame) -> pl.DataFrame:
    """
    Quita espacios al inicio/final de los nombres de columna. Los archivos
    de versiones anteriores a veces traían `VALOR OTROS ` o `COMENTARIOS `
    con espacios; aquí los normalizamos al nombre canónico sin espacios.
    """
    renombres = {c: c.strip() for c in df.columns if c != c.strip()}
    if renombres:
        df = df.rename(renombres)
    return df


def _leer_tabla_openpyxl(contenido_bytes: bytes, nombre_tabla: str) -> pl.DataFrame:
    """
    Fallback usando openpyxl. Busca la tabla por nombre en todas las hojas,
    obtiene su rango y devuelve un DataFrame.

    openpyxl entrega valores en sus tipos nativos de Python (datetime, int,
    float, str, None). Construimos el DataFrame columna a columna con dtype
    inferido para evitar depender de pyarrow.
    """
    try:
        from openpyxl import load_workbook
        from openpyxl.utils import range_boundaries
    except ImportError as exc:
        raise ImportError(
            "Se requiere openpyxl (o una version moderna de polars con fastexcel). "
            "Detalle: " + str(exc)
        )

    wb = load_workbook(io.BytesIO(contenido_bytes), data_only=True)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if nombre_tabla in ws.tables:
            tabla = ws.tables[nombre_tabla]
            ref = tabla.ref if hasattr(tabla, "ref") else tabla
            min_col, min_row, max_col, max_row = range_boundaries(ref)

            filas = list(ws.iter_rows(
                min_row=min_row, max_row=max_row,
                min_col=min_col, max_col=max_col,
                values_only=True,
            ))
            if not filas:
                raise ValueError("La tabla `" + nombre_tabla + "` está vacía.")

            # Strip de espacios para normalizar headers. Archivos antiguos
            # podían venir con "VALOR OTROS " o "COMENTARIOS " con espacios
            # al final; al hacer strip quedan en su forma canónica y la
            # validación contra los esquemas funciona.
            headers = [
                str(h).strip() if h is not None else "col_" + str(i)
                for i, h in enumerate(filas[0])
            ]
            cuerpo = filas[1:]

            series = []
            for i, header in enumerate(headers):
                valores = [r[i] if i < len(r) else None for r in cuerpo]
                series.append(_serie_polars(header, valores))
            return pl.DataFrame(series)

    raise ValueError(
        "No se encontro la tabla `" + nombre_tabla + "` en ninguna hoja del archivo. "
        "Verifica que el archivo tenga una tabla (objeto Table de Excel, "
        "no solo una hoja) con ese nombre exacto."
    )


def _serie_polars(nombre, valores):
    """
    Construye una pl.Series a partir de una lista de valores Python con
    Nones intercalados. Detecta el dtype inspeccionando los valores no nulos.
    """
    no_nulos = [v for v in valores if v is not None]
    if not no_nulos:
        return pl.Series(nombre, valores, dtype=pl.String)
    if all(isinstance(v, bool) for v in no_nulos):
        return pl.Series(nombre, valores, dtype=pl.Boolean)
    if all(isinstance(v, _dt.datetime) for v in no_nulos):
        return pl.Series(nombre, valores, dtype=pl.Datetime)
    if all(isinstance(v, _dt.date) for v in no_nulos):
        return pl.Series(nombre, valores, dtype=pl.Date)
    if all(isinstance(v, int) and not isinstance(v, bool) for v in no_nulos):
        return pl.Series(nombre, valores, dtype=pl.Int64)
    if all(isinstance(v, (int, float)) and not isinstance(v, bool) for v in no_nulos):
        return pl.Series(nombre, valores, dtype=pl.Float64)
    return pl.Series(
        nombre,
        [str(v) if v is not None else None for v in valores],
        dtype=pl.String,
    )


def normalizar_fecha(df: pl.DataFrame, columnas: list[str]) -> pl.DataFrame:
    """
    Convierte columnas de fecha de forma robusta independientemente del dtype
    con que Polars las haya leído.
    """
    exprs = []
    for col in columnas:
        if col not in df.columns:
            continue
        dtype = df.schema[col]
        if dtype == pl.Date:
            exprs.append(pl.col(col))
        elif dtype_es_fecha(dtype):
            exprs.append(pl.col(col).cast(pl.Date, strict=False))
        elif dtype in _TIPOS_NUMERICOS:
            exprs.append(
                pl.when(pl.col(col).is_not_null())
                .then(
                    pl.col(col).cast(pl.Int32, strict=False)
                    .map_elements(
                        lambda n: _dt.date(1899, 12, 30) + _dt.timedelta(days=int(n)),
                        return_dtype=pl.Date,
                    )
                )
                .otherwise(None)
                .alias(col)
            )
        elif dtype == pl.String:
            exprs.append(
                pl.when(pl.col(col).str.contains(r"^\d{2}/\d{2}/\d{4}$", literal=False))
                .then(pl.col(col).str.to_date("%d/%m/%Y", strict=False))
                .when(pl.col(col).str.contains(r"^\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2}$", literal=False))
                .then(pl.col(col).str.to_datetime("%Y-%m-%d %H:%M:%S", strict=False).cast(pl.Date))
                .otherwise(pl.col(col).str.to_date("%Y-%m-%d", strict=False))
                .alias(col)
            )
    if exprs:
        df = df.with_columns(exprs)
    return df


__all__ = [
    "ruta_reciente",
    "leer_excel_regalias",
    "leer_tabla_excel",
    "normalizar_fecha",
]
